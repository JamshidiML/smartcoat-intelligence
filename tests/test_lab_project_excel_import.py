from __future__ import annotations

import hashlib
import io
from collections.abc import Callable
from datetime import UTC, datetime

import pytest
from openpyxl import Workbook

from smartcoat.domain.lab_project_capture import (
    ApproachOutcome,
    AssessmentStatus,
    CaptureSourceKind,
    MeasurementState,
)
from smartcoat.domain.lab_project_capture import (
    TestOutcome as LabTestOutcome,
)
from smartcoat.services.lab_project_excel_import import (
    ExcelImportError,
    LabProjectExcelImporter,
    normalize_header,
)
from smartcoat.services.local_evidence_registry import LocalEvidenceDescriptor

NOW = datetime(2026, 8, 6, 10, 30, tzinfo=UTC)
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook_bytes(build: Callable[[Workbook], None]) -> bytes:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    build(workbook)
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _evidence(content: bytes) -> LocalEvidenceDescriptor:
    sha256 = hashlib.sha256(content).hexdigest()
    return LocalEvidenceDescriptor(
        evidence_id="asset:11111111-1111-4111-8111-111111111111",
        evidence_type="excel",
        original_filename="synthetic-projects.xlsx",
        media_type=XLSX_MEDIA_TYPE,
        sha256=sha256,
        source_reference=f"smartcoat-asset://synthetic-lab/{sha256}",
        captured_at=NOW,
        duplicate=False,
        size_bytes=len(content),
    )


def _import(content: bytes):
    return LabProjectExcelImporter().import_workbook(
        io.BytesIO(content),
        organization_id="synthetic-lab",
        evidence=_evidence(content),
    )


def test_english_headers_map_to_unconfirmed_candidate_with_cell_provenance() -> None:
    def build(workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Projects")
        sheet.append(
            [
                "Project Number",
                "Project Name",
                "Customer",
                "Request",
                "Goal",
                "Application",
                "Base Fabric",
                "Material",
                "Approach",
                "Result",
                "Test",
                "Test Result",
                "Temperature",
                "Sample",
                "Shipment Date",
                "Customer Feedback",
                "Production Feasibility",
                "Price",
                "Custom Lab Column",
            ]
        )
        sheet.append(
            [
                "P-SYN-001",
                "Synthetic heat shield",
                "Example Customer",
                "One-sided coated fabric",
                "Pass synthetic flame test",
                "High-temperature protection",
                "Synthetic glass fabric",
                "Synthetic filler",
                "Baseline",
                "Passed",
                "Flame test",
                "Passed",
                "210 degC",
                "S-02",
                "2026-08-05T12:00:00+00:00",
                "Pending",
                "Not assessed",
                "12.50",
                "preserved unknown value",
            ]
        )

    content = _workbook_bytes(build)
    first = _import(content)
    second = _import(content)

    assert first == second
    assert first.dry_run is True
    assert first.canonical_writes == 0
    assert len(first.candidates) == 1
    imported = first.candidates[0]
    candidate = imported.candidate
    assert candidate.source_kind is CaptureSourceKind.EXCEL
    assert candidate.source_language == "en"
    assert candidate.human_confirmed is False
    assert candidate.project.project_id == "P-SYN-001"
    assert candidate.project.customer_company == "Example Customer"
    assert candidate.substrate is not None
    assert candidate.substrate.substrate_name == "Synthetic glass fabric"
    assert candidate.materials[0].material_name == "Synthetic filler"
    assert candidate.approaches[0].outcome is ApproachOutcome.SUCCESSFUL
    assert candidate.approaches[0].production_feasibility_status is AssessmentStatus.NOT_ASSESSED
    assert candidate.process_parameters[0].numeric_value == 210
    assert candidate.process_parameters[0].unit == "degC"
    assert candidate.tests[0].outcome is LabTestOutcome.PASSED
    assert candidate.samples[0].sample_id == "S-02"
    assert candidate.evidence[0].source_reference == first.source_reference
    assert candidate.completeness_score < 100
    assert candidate.recommended_questions
    assert any(
        item.cell_reference == "A2" and item.target_field == "project_id"
        for item in imported.cell_provenance
    )
    assert imported.unmapped_values[0].display_value == "preserved unknown value"
    assert first.unmapped_columns[0].header == "Custom Lab Column"
    warning_codes = {item.code for item in first.warnings}
    assert "feedback_requires_review" in warning_codes
    assert "commercial_value_requires_currency" in warning_codes


def test_german_headers_and_explicit_not_measured_state() -> None:
    def build(workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Versuche")
        sheet.append(
            [
                "Projektnummer",
                "Projektname",
                "Kunde",
                "Anfrage",
                "Ziel",
                "Anwendung",
                "Grundgewebe",
                "Rohstoff",
                "Versuch",
                "Ergebnis",
                "Fehlergrund",
                "Prüfung",
                "Prüfergebnis",
                "Beschichtungsgewicht",
                "Probe",
                "Nachverfolgung",
                "Produktionsmachbarkeit",
                "Notizen",
            ]
        )
        sheet.append(
            [
                "P-SYN-DE-01",
                "Synthetischer Versuch",
                "Beispielkunde",
                "Flammschutz",
                "Prüfung bestehen",
                "Hitzeschutz",
                "Synthetisches Glasgewebe",
                "Synthetischer Füllstoff",
                "Ansatz eins",
                "fehlgeschlagen",
                "Haftung unzureichend",
                "Flammenprüfung",
                "nicht bestanden",
                "nicht gemessen",
                "S-01",
                "ausstehend",
                "nicht bewertet",
                "Synthetische Notiz",
            ]
        )

    report = _import(_workbook_bytes(build))
    candidate = report.candidates[0].candidate

    assert candidate.source_language == "de"
    assert candidate.approaches[0].outcome is ApproachOutcome.FAILED
    assert candidate.approaches[0].failure_reason == "Haftung unzureichend"
    assert candidate.tests[0].outcome is LabTestOutcome.FAILED
    assert candidate.process_parameters[0].measurement_state is MeasurementState.NOT_MEASURED
    assert "What was the exact coating weight?" in candidate.recommended_questions
    assert normalize_header("Prüfergebnis") == "prufergebnis"


def test_multiple_sheets_blank_rows_irregular_headers_and_row_error_isolation() -> None:
    def build(workbook: Workbook) -> None:
        first = workbook.create_sheet("Irregular")
        first.merge_cells("A1:B1")
        first["A1"] = "Synthetic workbook title"
        first.append([])
        first.append(["Project", "Approach", "Temperature"])
        first.append(["Valid project", "Valid approach", "180 degC"])
        first.append([None, None, None])
        first.append(["Invalid row", "Bad approach", "very hot"])

        second = workbook.create_sheet("Second")
        second.append(["Project Name", "Trial", "Result"])
        second.append(["Second project", "Second approach", "inconclusive"])

    report = _import(_workbook_bytes(build))

    assert report.sheet_names == ("Irregular", "Second")
    assert [(item.sheet_name, item.row_number) for item in report.candidates] == [
        ("Irregular", 4),
        ("Second", 2),
    ]
    assert len(report.row_errors) == 1
    assert report.row_errors[0].sheet_name == "Irregular"
    assert report.row_errors[0].row_number == 6
    assert report.row_errors[0].code == "invalid_numeric_parameter"
    warning_codes = {item.code for item in report.warnings}
    assert "merged_cells_present" in warning_codes
    assert "irregular_header_row" in warning_codes


def test_unknown_only_sheet_is_reported_without_candidate() -> None:
    def build(workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Unknown")
        sheet.append(["Mystery One", "Mystery Two"])
        sheet.append(["value", "value"])

    report = _import(_workbook_bytes(build))

    assert report.candidates == ()
    assert report.row_errors == ()
    assert [item.code for item in report.warnings] == ["no_recognized_headers"]


def test_invalid_xlsx_is_rejected() -> None:
    invalid = b"PK\x03\x04not-a-workbook"
    with pytest.raises(ExcelImportError, match="valid XLSX") as captured:
        _import(invalid)
    assert captured.value.code == "invalid_xlsx"


def test_import_rejects_cross_organization_evidence() -> None:
    def build(workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Projects")
        sheet.append(["Project Name"])
        sheet.append(["Synthetic project"])

    content = _workbook_bytes(build)
    with pytest.raises(ExcelImportError, match="requested organization") as captured:
        LabProjectExcelImporter().import_workbook(
            io.BytesIO(content),
            organization_id="other-lab",
            evidence=_evidence(content),
        )
    assert captured.value.code == "evidence_organization_mismatch"
