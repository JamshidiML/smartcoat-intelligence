"""Deterministic XLSX-to-candidate dry-run import for the intake pilot."""

from __future__ import annotations

import re
import unicodedata
from datetime import UTC, date, datetime
from enum import StrEnum
from typing import BinaryIO, Literal
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import AwareDatetime, BaseModel, ConfigDict

from smartcoat.domain.lab_project_capture import (
    ApproachOutcome,
    AssessmentStatus,
    FieldState,
    FollowUpStatus,
    LabProjectCaptureCandidate,
    MeasurementState,
    TestOutcome,
    apply_candidate_completeness,
)
from smartcoat.services.local_evidence_registry import (
    LocalEvidenceDescriptor,
    normalize_organization_id,
)

MAX_HEADER_SCAN_ROWS = 25
MAX_SHEETS = 32
MAX_ROWS_PER_SHEET = 10_000
MAX_COLUMNS_PER_SHEET = 256
MAX_CELL_TEXT = 4096
IMPORT_NAMESPACE = uuid5(NAMESPACE_URL, "urn:smartcoat:lab-project-xlsx-import:v1")
MAPPER_VERSION = "deterministic-xlsx-header-mapper-v1"


class ExcelImportError(ValueError):
    """A workbook-level failure that cannot be isolated to one row."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


class CanonicalColumn(StrEnum):
    PROJECT_ID = "project_id"
    PROJECT_NAME = "project_name"
    CUSTOMER_COMPANY = "customer_company"
    REQUEST_SUMMARY = "request_summary"
    SUCCESS_CRITERIA = "success_criteria"
    TARGET_APPLICATION = "target_application"
    SUBSTRATE_NAME = "substrate_name"
    MATERIAL_NAME = "material_name"
    FORMULATION = "formulation"
    APPROACH_TITLE = "approach_title"
    RESULT = "result"
    SUCCESS = "success"
    FAILURE = "failure"
    FAILURE_REASON = "failure_reason"
    TEST_NAME = "test_name"
    TEST_RESULT = "test_result"
    TEMPERATURE = "temperature"
    PRESSURE = "pressure"
    TIME = "time"
    SPEED = "speed"
    COATING_WEIGHT = "coating_weight"
    VISCOSITY = "viscosity"
    SAMPLE_ID = "sample_id"
    SHIPMENT_DATE = "shipment_date"
    CUSTOMER_FEEDBACK = "customer_feedback"
    FOLLOW_UP = "follow_up"
    PRODUCTION_FEASIBILITY = "production_feasibility"
    PRICE = "price"
    COST = "cost"
    NOTES = "notes"


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    decomposed = unicodedata.normalize("NFKD", str(value).strip().casefold())
    ascii_text = "".join(
        character for character in decomposed if not unicodedata.combining(character)
    )
    return " ".join(re.sub(r"[^a-z0-9]+", " ", ascii_text).split())


def _aliases(*values: str) -> tuple[str, ...]:
    return tuple(normalize_header(value) for value in values)


HEADER_ALIASES: dict[CanonicalColumn, tuple[str, ...]] = {
    CanonicalColumn.PROJECT_ID: _aliases(
        "project number", "project no", "project id", "projektnummer", "projekt nr", "projekt id"
    ),
    CanonicalColumn.PROJECT_NAME: _aliases("project", "project name", "projekt", "projektname"),
    CanonicalColumn.CUSTOMER_COMPANY: _aliases(
        "customer", "company", "customer company", "kunde", "firma", "unternehmen"
    ),
    CanonicalColumn.REQUEST_SUMMARY: _aliases(
        "request", "customer request", "anfrage", "kundenanfrage", "anforderung"
    ),
    CanonicalColumn.SUCCESS_CRITERIA: _aliases(
        "goal", "success criteria", "target", "ziel", "erfolgskriterien"
    ),
    CanonicalColumn.TARGET_APPLICATION: _aliases(
        "application", "target application", "anwendung", "einsatzgebiet"
    ),
    CanonicalColumn.SUBSTRATE_NAME: _aliases(
        "base fabric", "substrate", "base substrate", "grundgewebe", "basisgewebe", "substrat"
    ),
    CanonicalColumn.MATERIAL_NAME: _aliases(
        "material", "raw material", "raw materials", "rohmaterial", "rohstoff", "rohstoffe"
    ),
    CanonicalColumn.FORMULATION: _aliases(
        "formulation", "recipe", "formulierung", "rezept", "rezeptur"
    ),
    CanonicalColumn.APPROACH_TITLE: _aliases(
        "approach", "trial", "experimental approach", "ansatz", "versuch", "versuchsansatz"
    ),
    CanonicalColumn.RESULT: _aliases("result", "outcome", "ergebnis", "resultat"),
    CanonicalColumn.SUCCESS: _aliases("success", "successful", "erfolg", "erfolgreich"),
    CanonicalColumn.FAILURE: _aliases(
        "failure", "failed", "fehler", "fehlgeschlagen", "misserfolg"
    ),
    CanonicalColumn.FAILURE_REASON: _aliases(
        "failure reason", "reason for failure", "fehlergrund", "grund des scheiterns"
    ),
    CanonicalColumn.TEST_NAME: _aliases(
        "test", "test method", "prufung", "prufmethode", "pruefung"
    ),
    CanonicalColumn.TEST_RESULT: _aliases(
        "test result", "test outcome", "prufergebnis", "pruefergebnis", "testergebnis"
    ),
    CanonicalColumn.TEMPERATURE: _aliases(
        "temperature", "curing temperature", "temperatur", "hartetemperatur"
    ),
    CanonicalColumn.PRESSURE: _aliases("pressure", "druck"),
    CanonicalColumn.TIME: _aliases("time", "duration", "curing time", "zeit", "dauer", "hartezeit"),
    CanonicalColumn.SPEED: _aliases(
        "speed", "line speed", "geschwindigkeit", "liniengeschwindigkeit"
    ),
    CanonicalColumn.COATING_WEIGHT: _aliases(
        "coating weight", "coat weight", "beschichtungsgewicht", "auflage", "warenauflage"
    ),
    CanonicalColumn.VISCOSITY: _aliases("viscosity", "viskositat", "viskositaet"),
    CanonicalColumn.SAMPLE_ID: _aliases(
        "sample", "sample id", "probe", "proben id", "probennummer"
    ),
    CanonicalColumn.SHIPMENT_DATE: _aliases(
        "shipment date", "sent date", "versanddatum", "versendet am"
    ),
    CanonicalColumn.CUSTOMER_FEEDBACK: _aliases(
        "customer feedback", "feedback", "kundenfeedback", "kundenruckmeldung", "kundenrueckmeldung"
    ),
    CanonicalColumn.FOLLOW_UP: _aliases(
        "follow-up", "follow up", "followup", "nachverfolgung", "nachfassen"
    ),
    CanonicalColumn.PRODUCTION_FEASIBILITY: _aliases(
        "production feasibility",
        "manufacturing feasibility",
        "produktionsmachbarkeit",
        "serienmachbarkeit",
    ),
    CanonicalColumn.PRICE: _aliases("price", "preis"),
    CanonicalColumn.COST: _aliases("cost", "costs", "kosten"),
    CanonicalColumn.NOTES: _aliases(
        "comment", "comments", "notes", "kommentar", "notiz", "notizen"
    ),
}
ALIAS_TO_COLUMN = {alias: column for column, aliases in HEADER_ALIASES.items() for alias in aliases}
GERMAN_ALIASES = {
    normalize_header(value)
    for value in (
        "projektnummer",
        "projekt",
        "projektname",
        "kunde",
        "firma",
        "anfrage",
        "ziel",
        "anwendung",
        "grundgewebe",
        "substrat",
        "rohstoff",
        "formulierung",
        "rezeptur",
        "ansatz",
        "versuch",
        "ergebnis",
        "erfolg",
        "fehler",
        "fehlergrund",
        "pruefung",
        "pruefergebnis",
        "temperatur",
        "druck",
        "zeit",
        "geschwindigkeit",
        "beschichtungsgewicht",
        "viskositaet",
        "probe",
        "versanddatum",
        "kundenfeedback",
        "nachverfolgung",
        "produktionsmachbarkeit",
        "preis",
        "kosten",
        "kommentar",
        "notizen",
    )
}


class ImportWarning(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    code: str
    message: str
    sheet_name: str | None = None
    row_number: int | None = None
    cell_reference: str | None = None


class RowImportError(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    sheet_name: str
    row_number: int
    code: str
    message: str
    cell_reference: str | None = None


class CellProvenance(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    target_field: str
    workbook_source_reference: str
    sheet_name: str
    row_number: int
    column_number: int
    cell_reference: str
    header: str
    display_value: str


class UnmappedColumn(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    sheet_name: str
    header_row: int
    column_number: int
    cell_reference: str
    header: str
    reason: str = "unknown_header"


class UnmappedCellValue(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    sheet_name: str
    row_number: int
    cell_reference: str
    header: str
    display_value: str
    reason: str


class ImportedCandidate(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    sheet_name: str
    row_number: int
    candidate: LabProjectCaptureCandidate
    cell_provenance: tuple[CellProvenance, ...]
    unmapped_values: tuple[UnmappedCellValue, ...]


class WorkbookImportReport(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    dry_run: Literal[True] = True
    canonical_writes: Literal[0] = 0
    organization_id: str
    original_filename: str
    source_reference: str
    sha256: str
    sheet_names: tuple[str, ...]
    candidates: tuple[ImportedCandidate, ...]
    unmapped_columns: tuple[UnmappedColumn, ...]
    row_errors: tuple[RowImportError, ...]
    warnings: tuple[ImportWarning, ...]


class _RowValueError(ValueError):
    def __init__(self, code: str, message: str, cell_reference: str | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.cell_reference = cell_reference


class LabProjectExcelImporter:
    """Convert workbook rows into review-only candidates without persistence."""

    def import_workbook(
        self,
        stream: BinaryIO,
        *,
        organization_id: str,
        evidence: LocalEvidenceDescriptor,
    ) -> WorkbookImportReport:
        organization = normalize_organization_id(organization_id)
        if (
            evidence.evidence_type.value != "excel"
            or evidence.media_type
            != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            raise ExcelImportError("invalid_evidence_type", "Excel import requires XLSX evidence")
        if not evidence.source_reference.startswith(f"smartcoat-asset://{organization}/"):
            raise ExcelImportError(
                "evidence_organization_mismatch",
                "Evidence descriptor does not belong to the requested organization",
            )
        try:
            stream.seek(0)
            workbook = load_workbook(
                stream,
                data_only=True,
                read_only=False,
                keep_links=False,
            )
        except (InvalidFileException, OSError, ValueError, KeyError) as error:
            raise ExcelImportError("invalid_xlsx", "Workbook is not a valid XLSX file") from error
        except Exception as error:
            if error.__class__.__module__.startswith(("zipfile", "xml")):
                raise ExcelImportError(
                    "invalid_xlsx", "Workbook is not a valid XLSX file"
                ) from error
            raise

        if len(workbook.worksheets) > MAX_SHEETS:
            raise ExcelImportError("workbook_too_large", "Workbook exceeds the sheet limit")

        candidates: list[ImportedCandidate] = []
        unmapped_columns: list[UnmappedColumn] = []
        row_errors: list[RowImportError] = []
        warnings: list[ImportWarning] = []

        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            if (
                worksheet.max_row > MAX_ROWS_PER_SHEET
                or worksheet.max_column > MAX_COLUMNS_PER_SHEET
            ):
                warnings.append(
                    ImportWarning(
                        code="sheet_limit_exceeded",
                        message=(
                            "Sheet was skipped because its declared dimensions exceed pilot limits"
                        ),
                        sheet_name=worksheet.title,
                    )
                )
                continue
            if worksheet.merged_cells.ranges:
                warnings.append(
                    ImportWarning(
                        code="merged_cells_present",
                        message="Merged cells may make row interpretation ambiguous",
                        sheet_name=worksheet.title,
                    )
                )

            header_row = self._find_header_row(worksheet)
            if header_row is None:
                warnings.append(
                    ImportWarning(
                        code="no_recognized_headers",
                        message="Sheet contains no recognized German or English intake headers",
                        sheet_name=worksheet.title,
                    )
                )
                continue
            if header_row != 1:
                warnings.append(
                    ImportWarning(
                        code="irregular_header_row",
                        message=f"Headers were detected on row {header_row} instead of row 1",
                        sheet_name=worksheet.title,
                        row_number=header_row,
                    )
                )

            mapped, sheet_unmapped, mapping_warnings = self._map_headers(worksheet, header_row)
            unmapped_columns.extend(sheet_unmapped)
            warnings.extend(mapping_warnings)
            source_language = self._source_language(worksheet, header_row)

            for row_number in range(header_row + 1, worksheet.max_row + 1):
                if self._row_is_blank(worksheet, row_number):
                    continue
                try:
                    imported, row_warnings = self._build_candidate(
                        worksheet,
                        sheet_index=sheet_index,
                        header_row=header_row,
                        row_number=row_number,
                        mapped=mapped,
                        unmapped_columns=sheet_unmapped,
                        source_language=source_language,
                        organization_id=organization,
                        evidence=evidence,
                    )
                    if imported is None:
                        warnings.append(
                            ImportWarning(
                                code="row_without_mapped_values",
                                message=(
                                    "Non-blank row contained no values under recognized headers"
                                ),
                                sheet_name=worksheet.title,
                                row_number=row_number,
                            )
                        )
                        continue
                    candidates.append(imported)
                    warnings.extend(row_warnings)
                except _RowValueError as error:
                    row_errors.append(
                        RowImportError(
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            code=error.code,
                            message=str(error),
                            cell_reference=error.cell_reference,
                        )
                    )
                except (TypeError, ValueError) as error:
                    row_errors.append(
                        RowImportError(
                            sheet_name=worksheet.title,
                            row_number=row_number,
                            code="invalid_candidate_row",
                            message=str(error),
                        )
                    )

        sheet_names = tuple(workbook.sheetnames)
        workbook.close()
        return WorkbookImportReport(
            organization_id=organization,
            original_filename=evidence.original_filename,
            source_reference=evidence.source_reference,
            sha256=evidence.sha256,
            sheet_names=sheet_names,
            candidates=tuple(candidates),
            unmapped_columns=tuple(unmapped_columns),
            row_errors=tuple(row_errors),
            warnings=tuple(warnings),
        )

    @staticmethod
    def _find_header_row(worksheet: Worksheet) -> int | None:
        best_row: int | None = None
        best_score = 0
        max_row = min(worksheet.max_row, MAX_HEADER_SCAN_ROWS)
        max_column = min(worksheet.max_column, MAX_COLUMNS_PER_SHEET)
        for row_number in range(1, max_row + 1):
            matches = {
                ALIAS_TO_COLUMN[normalized]
                for column_number in range(1, max_column + 1)
                if (normalized := normalize_header(worksheet.cell(row_number, column_number).value))
                in ALIAS_TO_COLUMN
            }
            if len(matches) > best_score:
                best_row = row_number
                best_score = len(matches)
        return best_row

    @staticmethod
    def _map_headers(
        worksheet: Worksheet,
        header_row: int,
    ) -> tuple[dict[CanonicalColumn, int], list[UnmappedColumn], list[ImportWarning]]:
        mapped: dict[CanonicalColumn, int] = {}
        unmapped: list[UnmappedColumn] = []
        warnings: list[ImportWarning] = []
        for column_number in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(header_row, column_number)
            display = _display_value(cell.value)
            if not display:
                continue
            canonical = ALIAS_TO_COLUMN.get(normalize_header(display))
            if canonical is None:
                unmapped.append(
                    UnmappedColumn(
                        sheet_name=worksheet.title,
                        header_row=header_row,
                        column_number=column_number,
                        cell_reference=cell.coordinate,
                        header=display,
                    )
                )
            elif canonical in mapped:
                unmapped.append(
                    UnmappedColumn(
                        sheet_name=worksheet.title,
                        header_row=header_row,
                        column_number=column_number,
                        cell_reference=cell.coordinate,
                        header=display,
                        reason="duplicate_mapping",
                    )
                )
                warnings.append(
                    ImportWarning(
                        code="ambiguous_duplicate_header",
                        message=f"Multiple columns map to {canonical.value}; the first was used",
                        sheet_name=worksheet.title,
                        cell_reference=cell.coordinate,
                    )
                )
            else:
                mapped[canonical] = column_number
        return mapped, unmapped, warnings

    @staticmethod
    def _source_language(worksheet: Worksheet, header_row: int) -> str:
        normalized = {
            normalize_header(worksheet.cell(header_row, column_number).value)
            for column_number in range(1, worksheet.max_column + 1)
        }
        return "de" if normalized & GERMAN_ALIASES else "en"

    @staticmethod
    def _row_is_blank(worksheet: Worksheet, row_number: int) -> bool:
        return all(
            _display_value(worksheet.cell(row_number, column_number).value) == ""
            for column_number in range(1, worksheet.max_column + 1)
        )

    def _build_candidate(
        self,
        worksheet: Worksheet,
        *,
        sheet_index: int,
        header_row: int,
        row_number: int,
        mapped: dict[CanonicalColumn, int],
        unmapped_columns: list[UnmappedColumn],
        source_language: str,
        organization_id: str,
        evidence: LocalEvidenceDescriptor,
    ) -> tuple[ImportedCandidate | None, list[ImportWarning]]:
        values: dict[CanonicalColumn, object] = {}
        provenance: list[CellProvenance] = []
        row_warnings: list[ImportWarning] = []
        for field, column_number in mapped.items():
            cell = worksheet.cell(row_number, column_number)
            display = _display_value(cell.value)
            if display == "":
                continue
            if len(display) > MAX_CELL_TEXT:
                raise _RowValueError(
                    "cell_value_too_large",
                    "Cell value exceeds the pilot text limit",
                    cell.coordinate,
                )
            values[field] = cell.value
            header = _display_value(worksheet.cell(header_row, column_number).value)
            provenance.append(
                CellProvenance(
                    target_field=field.value,
                    workbook_source_reference=evidence.source_reference,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    column_number=column_number,
                    cell_reference=cell.coordinate,
                    header=header,
                    display_value=display,
                )
            )
        if not values:
            return None, row_warnings

        unmapped_values = self._unmapped_values(worksheet, row_number, unmapped_columns)
        field_states: dict[str, FieldState] = {}

        def text_value(field: CanonicalColumn, path: str) -> str | None:
            raw = values.get(field)
            if raw is None:
                return None
            display = _display_value(raw)
            state = _explicit_state(display)
            if state is not None:
                field_states[path] = state
                return None
            return display

        project: dict[str, object] = {
            "project_id": text_value(CanonicalColumn.PROJECT_ID, "project.project_id"),
            "project_name": text_value(CanonicalColumn.PROJECT_NAME, "project.project_name"),
            "customer_company": text_value(
                CanonicalColumn.CUSTOMER_COMPANY, "project.customer_company"
            ),
            "request_summary": text_value(
                CanonicalColumn.REQUEST_SUMMARY, "project.request_summary"
            ),
            "target_application": text_value(
                CanonicalColumn.TARGET_APPLICATION, "project.target_application"
            ),
        }
        success_criterion = text_value(CanonicalColumn.SUCCESS_CRITERIA, "project.success_criteria")
        if success_criterion is not None:
            project["success_criteria"] = (success_criterion,)
        project = {key: value for key, value in project.items() if value is not None}

        substrate_name = text_value(CanonicalColumn.SUBSTRATE_NAME, "substrate.substrate_name")
        substrate = {"substrate_name": substrate_name} if substrate_name is not None else None

        row_key = f"{evidence.sha256}:{worksheet.title}:{row_number}"
        approach_id = f"A-{sheet_index:02d}-{row_number:05d}"
        material_text = text_value(CanonicalColumn.MATERIAL_NAME, "materials.0.material_name")
        formulation_text = text_value(CanonicalColumn.FORMULATION, "materials.0.material_name")
        materials: list[dict[str, object]] = []
        selected_material = material_text or formulation_text
        if selected_material is not None:
            materials.append(
                {
                    "material_id": f"M-{sheet_index:02d}-{row_number:05d}",
                    "material_name": selected_material,
                }
            )
            if formulation_text is not None and material_text is None:
                row_warnings.append(
                    self._warning(
                        worksheet,
                        row_number,
                        "formulation_preserved_as_material_text",
                        "Formulation/recipe text was preserved as a reviewable material entry",
                    )
                )

        approach_title = text_value(
            CanonicalColumn.APPROACH_TITLE, f"approaches.{approach_id}.title"
        )
        result = text_value(CanonicalColumn.RESULT, f"approaches.{approach_id}.outcome_summary")
        failure_reason = text_value(
            CanonicalColumn.FAILURE_REASON, f"approaches.{approach_id}.failure_reason"
        )
        notes = text_value(CanonicalColumn.NOTES, f"approaches.{approach_id}.description")
        has_approach_detail = any(
            field in values
            for field in {
                CanonicalColumn.APPROACH_TITLE,
                CanonicalColumn.RESULT,
                CanonicalColumn.SUCCESS,
                CanonicalColumn.FAILURE,
                CanonicalColumn.FAILURE_REASON,
                CanonicalColumn.TEST_NAME,
                CanonicalColumn.TEST_RESULT,
                CanonicalColumn.TEMPERATURE,
                CanonicalColumn.PRESSURE,
                CanonicalColumn.TIME,
                CanonicalColumn.SPEED,
                CanonicalColumn.COATING_WEIGHT,
                CanonicalColumn.VISCOSITY,
                CanonicalColumn.SAMPLE_ID,
                CanonicalColumn.NOTES,
                CanonicalColumn.PRODUCTION_FEASIBILITY,
            }
        )
        approaches: list[dict[str, object]] = []
        if has_approach_detail:
            outcome = self._outcome(values, result, failure_reason)
            feasibility, feasibility_notes = self._assessment(
                values.get(CanonicalColumn.PRODUCTION_FEASIBILITY)
            )
            approach: dict[str, object] = {
                "approach_id": approach_id,
                "outcome": outcome,
            }
            if approach_title is not None:
                approach["title"] = approach_title
            if result is not None:
                approach["outcome_summary"] = result
            if failure_reason is not None:
                approach["failure_reason"] = failure_reason
            if notes is not None:
                approach["description"] = notes
            if feasibility is not None:
                approach["production_feasibility_status"] = feasibility
            if feasibility_notes is not None:
                approach["production_feasibility_notes"] = feasibility_notes
            approaches.append(approach)

        parameters: list[dict[str, object]] = []
        for field, name in (
            (CanonicalColumn.TEMPERATURE, "temperature"),
            (CanonicalColumn.PRESSURE, "pressure"),
            (CanonicalColumn.TIME, "time"),
            (CanonicalColumn.SPEED, "speed"),
            (CanonicalColumn.COATING_WEIGHT, "coating weight"),
            (CanonicalColumn.VISCOSITY, "viscosity"),
        ):
            if field not in values:
                continue
            column_number = mapped[field]
            parameters.append(
                self._process_parameter(
                    values[field],
                    approach_id=approach_id,
                    parameter_name=name,
                    cell_reference=worksheet.cell(row_number, column_number).coordinate,
                    warnings=row_warnings,
                    worksheet=worksheet,
                    row_number=row_number,
                )
            )

        tests: list[dict[str, object]] = []
        test_name = text_value(CanonicalColumn.TEST_NAME, f"tests.{approach_id}.test_name")
        test_result = text_value(CanonicalColumn.TEST_RESULT, f"tests.{approach_id}.result")
        if test_name is not None:
            test: dict[str, object] = {
                "approach_id": approach_id,
                "test_name": test_name,
                "outcome": self._test_outcome(test_result),
            }
            if test_result is not None:
                test["text_result"] = test_result
            tests.append(test)
        elif test_result is not None:
            row_warnings.append(
                self._warning(
                    worksheet,
                    row_number,
                    "test_result_without_test_name",
                    (
                        "Test result was preserved in provenance but requires a test name before "
                        "mapping"
                    ),
                )
            )

        samples: list[dict[str, object]] = []
        sample_id = text_value(CanonicalColumn.SAMPLE_ID, "samples.0.sample_id")
        if sample_id is not None:
            sample: dict[str, object] = {
                "sample_id": sample_id,
                "approach_id": approach_id,
                "physical_archive_status": "unknown",
            }
            if CanonicalColumn.SHIPMENT_DATE in values:
                shipment_cell = worksheet.cell(row_number, mapped[CanonicalColumn.SHIPMENT_DATE])
                sample["sent_at"] = self._parse_datetime(
                    values[CanonicalColumn.SHIPMENT_DATE],
                    worksheet=worksheet,
                    row_number=row_number,
                    cell_reference=shipment_cell.coordinate,
                    warnings=row_warnings,
                )
            follow_up = self._follow_up_status(values.get(CanonicalColumn.FOLLOW_UP))
            if follow_up is not None:
                sample["follow_up_status"] = follow_up
            samples.append(sample)
        elif CanonicalColumn.SHIPMENT_DATE in values:
            row_warnings.append(
                self._warning(
                    worksheet,
                    row_number,
                    "shipment_without_sample",
                    (
                        "Shipment date was preserved in provenance but requires a sample ID before "
                        "mapping"
                    ),
                )
            )

        if CanonicalColumn.CUSTOMER_FEEDBACK in values:
            row_warnings.append(
                self._warning(
                    worksheet,
                    row_number,
                    "feedback_requires_review",
                    (
                        "Customer feedback was preserved in provenance; sender and received date "
                        "are required"
                    ),
                )
            )
        for cost_field in (CanonicalColumn.PRICE, CanonicalColumn.COST):
            if cost_field in values:
                row_warnings.append(
                    self._warning(
                        worksheet,
                        row_number,
                        "commercial_value_requires_currency",
                        (
                            f"{cost_field.value.title()} was preserved in provenance and requires "
                            "currency/basis review"
                        ),
                    )
                )

        candidate_warnings = tuple(dict.fromkeys(item.message for item in row_warnings))
        candidate = LabProjectCaptureCandidate.model_validate(
            {
                "capture_session_id": uuid5(IMPORT_NAMESPACE, f"{organization_id}:{row_key}"),
                "source_kind": "excel",
                "source_language": source_language,
                "extraction_model": MAPPER_VERSION,
                "extraction_started_at": evidence.captured_at,
                "extraction_completed_at": evidence.captured_at,
                "project": project,
                "substrate": substrate,
                "materials": materials,
                "approaches": approaches,
                "process_parameters": parameters,
                "tests": tests,
                "samples": samples,
                "evidence": (evidence.as_candidate_evidence(),),
                "field_states": field_states,
                "extraction_warnings": candidate_warnings,
                "human_confirmed": False,
            }
        )
        candidate = apply_candidate_completeness(candidate)
        return (
            ImportedCandidate(
                sheet_name=worksheet.title,
                row_number=row_number,
                candidate=candidate,
                cell_provenance=tuple(provenance),
                unmapped_values=tuple(unmapped_values),
            ),
            row_warnings,
        )

    @staticmethod
    def _unmapped_values(
        worksheet: Worksheet,
        row_number: int,
        columns: list[UnmappedColumn],
    ) -> list[UnmappedCellValue]:
        values: list[UnmappedCellValue] = []
        for column in columns:
            cell = worksheet.cell(row_number, column.column_number)
            display = _display_value(cell.value)
            if display:
                values.append(
                    UnmappedCellValue(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        cell_reference=cell.coordinate,
                        header=column.header,
                        display_value=display[:MAX_CELL_TEXT],
                        reason=column.reason,
                    )
                )
        return values

    @staticmethod
    def _outcome(
        values: dict[CanonicalColumn, object],
        result: str | None,
        failure_reason: str | None,
    ) -> ApproachOutcome:
        failure = _display_value(values.get(CanonicalColumn.FAILURE)).casefold()
        success = _display_value(values.get(CanonicalColumn.SUCCESS)).casefold()
        result_normalized = (result or "").casefold()
        if (
            failure_reason
            or _is_truthy(failure)
            or _contains_any(
                result_normalized, "failed", "failure", "fail", "fehlgeschlagen", "gescheitert"
            )
        ):
            return ApproachOutcome.FAILED
        if _is_truthy(success) or _contains_any(
            result_normalized, "successful", "passed", "success", "bestanden", "erfolgreich"
        ):
            return ApproachOutcome.SUCCESSFUL
        if _contains_any(result_normalized, "partial", "teilweise"):
            return ApproachOutcome.PARTIALLY_SUCCESSFUL
        if result:
            return ApproachOutcome.INCONCLUSIVE
        return ApproachOutcome.PLANNED

    @staticmethod
    def _assessment(value: object | None) -> tuple[AssessmentStatus | None, str | None]:
        display = _display_value(value)
        if not display:
            return None, None
        normalized = display.casefold()
        if _explicit_state(display) in {FieldState.UNKNOWN, FieldState.MISSING}:
            return AssessmentStatus.UNKNOWN, None
        if _contains_any(normalized, "not applicable", "n/a", "nicht zutreffend"):
            return AssessmentStatus.NOT_APPLICABLE, None
        if _contains_any(normalized, "not assessed", "not evaluated", "nicht bewertet", "nein"):
            return AssessmentStatus.NOT_ASSESSED, None
        if _is_truthy(normalized) or _contains_any(
            normalized, "assessed", "evaluated", "feasible", "bewertet", "machbar"
        ):
            return AssessmentStatus.ASSESSED, display
        return AssessmentStatus.UNKNOWN, display

    @staticmethod
    def _process_parameter(
        value: object,
        *,
        approach_id: str,
        parameter_name: str,
        cell_reference: str,
        warnings: list[ImportWarning],
        worksheet: Worksheet,
        row_number: int,
    ) -> dict[str, object]:
        display = _display_value(value)
        state = _explicit_measurement_state(display)
        base: dict[str, object] = {
            "approach_id": approach_id,
            "process_stage": "imported",
            "parameter_name": parameter_name,
        }
        if state is not None:
            base["measurement_state"] = state
            return base

        numeric, unit = _numeric_and_unit(value)
        if numeric is not None and unit is not None:
            base.update(
                {
                    "numeric_value": numeric,
                    "unit": unit,
                    "measurement_state": MeasurementState.KNOWN,
                }
            )
            return base
        if numeric is not None:
            base.update(
                {
                    "text_value": display,
                    "measurement_state": MeasurementState.KNOWN,
                    "source_note": "Numeric source value has no declared unit",
                }
            )
            warnings.append(
                ImportWarning(
                    code="parameter_unit_missing",
                    message=f"{parameter_name.title()} requires unit confirmation",
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    cell_reference=cell_reference,
                )
            )
            return base
        if re.search(r"\d", display):
            raise _RowValueError(
                "invalid_numeric_parameter",
                f"Could not parse {parameter_name} as a deterministic numeric value",
                cell_reference,
            )
        raise _RowValueError(
            "invalid_numeric_parameter",
            f"{parameter_name.title()} must be numeric or an explicit unknown state",
            cell_reference,
        )

    @staticmethod
    def _test_outcome(result: str | None) -> TestOutcome:
        if result is None:
            return TestOutcome.NOT_MEASURED
        normalized = result.casefold()
        if _contains_any(normalized, "not measured", "nicht gemessen"):
            return TestOutcome.NOT_MEASURED
        if _contains_any(normalized, "failed", "fail", "nicht bestanden", "fehlgeschlagen"):
            return TestOutcome.FAILED
        if _contains_any(normalized, "partial", "teilweise"):
            return TestOutcome.PARTIALLY_PASSED
        if _contains_any(normalized, "passed", "pass", "bestanden", "erfolgreich"):
            return TestOutcome.PASSED
        return TestOutcome.INCONCLUSIVE

    @staticmethod
    def _follow_up_status(value: object | None) -> FollowUpStatus | None:
        display = _display_value(value)
        if not display:
            return None
        normalized = display.casefold()
        if _contains_any(normalized, "not required", "nicht erforderlich", "n/a"):
            return FollowUpStatus.NOT_REQUIRED
        if _contains_any(
            normalized, "feedback received", "ruckmeldung erhalten", "rueckmeldung erhalten"
        ):
            return FollowUpStatus.FEEDBACK_RECEIVED
        if _contains_any(normalized, "contacted", "kontaktiert"):
            return FollowUpStatus.CONTACTED
        if _contains_any(normalized, "closed", "abgeschlossen"):
            return FollowUpStatus.CLOSED
        if _contains_any(normalized, "overdue", "uberfallig", "ueberfaellig"):
            return FollowUpStatus.OVERDUE
        return FollowUpStatus.PENDING

    @staticmethod
    def _parse_datetime(
        value: object,
        *,
        worksheet: Worksheet,
        row_number: int,
        cell_reference: str,
        warnings: list[ImportWarning],
    ) -> AwareDatetime:
        parsed: datetime
        assumed_timezone = False
        if isinstance(value, datetime):
            parsed = value
        elif isinstance(value, date):
            parsed = datetime.combine(value, datetime.min.time())
        else:
            try:
                parsed = datetime.fromisoformat(_display_value(value))
            except ValueError as error:
                raise _RowValueError(
                    "invalid_shipment_date",
                    "Shipment date must be an Excel date or ISO-8601 value",
                    cell_reference,
                ) from error
        if parsed.tzinfo is None or parsed.utcoffset() is None:
            parsed = parsed.replace(tzinfo=UTC)
            assumed_timezone = True
        if assumed_timezone:
            warnings.append(
                ImportWarning(
                    code="shipment_timezone_assumed_utc",
                    message="Shipment date had no timezone; UTC was applied for review",
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    cell_reference=cell_reference,
                )
            )
        return parsed.astimezone(UTC)

    @staticmethod
    def _warning(
        worksheet: Worksheet,
        row_number: int,
        code: str,
        message: str,
    ) -> ImportWarning:
        return ImportWarning(
            code=code,
            message=message,
            sheet_name=worksheet.title,
            row_number=row_number,
        )


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _contains_any(value: str, *terms: str) -> bool:
    return any(term in value for term in terms)


def _is_truthy(value: str) -> bool:
    return normalize_header(value) in {
        "1",
        "true",
        "yes",
        "ja",
        "y",
        "x",
        "successful",
        "erfolgreich",
    }


def _explicit_state(value: str) -> FieldState | None:
    normalized = normalize_header(value)
    states = {
        "unknown": FieldState.UNKNOWN,
        "unbekannt": FieldState.UNKNOWN,
        "missing": FieldState.MISSING,
        "fehlt": FieldState.MISSING,
        "not measured": FieldState.NOT_MEASURED,
        "nicht gemessen": FieldState.NOT_MEASURED,
        "not applicable": FieldState.NOT_APPLICABLE,
        "nicht zutreffend": FieldState.NOT_APPLICABLE,
        "conflicting": FieldState.CONFLICTING,
        "widerspruchlich": FieldState.CONFLICTING,
        "widerspruechlich": FieldState.CONFLICTING,
    }
    return states.get(normalized)


def _explicit_measurement_state(value: str) -> MeasurementState | None:
    state = _explicit_state(value)
    if state is None or state is FieldState.MISSING:
        return MeasurementState.UNKNOWN if state is FieldState.MISSING else None
    return MeasurementState(state.value)


def _numeric_and_unit(value: object) -> tuple[float | None, str | None]:
    if isinstance(value, bool):
        return None, None
    if isinstance(value, int | float):
        return float(value), None
    display = _display_value(value)
    match = re.fullmatch(r"\s*([-+]?\d+(?:[.,]\d+)?)\s*([^\d\s].*?)?\s*", display)
    if match is None:
        return None, None
    numeric = float(match.group(1).replace(",", "."))
    unit = match.group(2).strip() if match.group(2) else None
    return numeric, unit


__all__ = [
    "CellProvenance",
    "ExcelImportError",
    "ImportWarning",
    "ImportedCandidate",
    "LabProjectExcelImporter",
    "RowImportError",
    "UnmappedColumn",
    "UnmappedCellValue",
    "WorkbookImportReport",
    "normalize_header",
]
